# !/usr/bin/env python
# -*- coding:utf-8 -*-
# author:ouyang time:2021/11/19

import requests
from bs4 import BeautifulSoup as bs
import re
import openpyxl
import os
import json
"""
爬取的内容示例1：
           11月19日，宁夏应对新型冠状病毒肺炎疫情工作指挥部发布最新全区新冠肺炎疫情通报。1月18日0时-24时，全区报告无新增新冠肺炎本土确诊病例、疑似病例和无症状感染者，'
           '新增治愈出院2例。截至18日24时，现有确诊病例15例。自10月17日以来，全区累计报告确诊病例45例，累计治愈出院30例。
      示例2：10月20日0时-24时，全区报告新增新冠肺炎确诊病例4例（银川2例，吴忠2例），无新增疑似病例、无症状感染者。截至20日24时，
           现有确诊病例6例，无现有疑似病例、无症状感染者。全区累计报告确诊病例83例(含4例境外输入病例)。
      总网页：
           http://wsjkw.nx.gov.cn/sy_279/yqfkdt/yqsd/index.html
      分网页：    
           http://wsjkw.nx.gov.cn/sy_279/yqfkdt/yqsd/index_1.html    
           http://wsjkw.nx.gov.cn/sy_279/yqfkdt/yqsd/index_2.html
           
程序设计框架：
        1：录入宁夏自治区的所有省市，并初始化表格；
           表格格式：
           sheet0: 日期  省的确诊增加人数  该省迁入人数  该省迁出人数 
           sheet*: 日期   市的确诊数据     市迁入数据   市迁出数据    迁入迁出数据总和
           （注意：因为网页的制作时间会晚一天，因此最后统计的时间和人数要错开一天）
        2：通过宁夏卫健委官网爬取从四月1日起到现在的每天的公布的数据，找到每天具体地区的确诊人员；
           步骤：通过总网->查找分网->爬取关键帧数据
        3：获取4月1日来的迁徙数据，并记录表：
        4：分析数据：
"""
""" ****************************************************第一步，表格的准备工作**************"""

# 手动录入, 并分别新建城市的字典，建立时间-新增人数键值对
city_names = ['银川', '固原', '石嘴山', '吴忠', '中卫']
Yinchuan = {}
Guyuan   = {}
Shizuishan = {}
Wuzhong  = {}
Zhongwei = {}
Total = {}
# 时间列表  网页
html_times = []
html_names = []


# 设置表格的路径位置
path = r"F:\important\研究生日常\作业\数据挖掘\第二次实验"
os.chdir(path)  # 修改工作路径
workbook = openpyxl.load_workbook('data.xlsx')


# 获取已有活动表，即全省的表
privacy = workbook['全省']
city_sheet_yic = workbook['银川']  # 获取新的表的对象
city_sheet_guy = workbook['固原']  # 获取新的表的对象
city_sheet_shz = workbook['石嘴山']  # 获取新的表的对象
city_sheet_wuz = workbook['吴忠']  # 获取新的表的对象
city_sheet_zhw = workbook['中卫']  # 获取新的表的对象


print("1-正在保存文件...")
# 这里需要保存一次，不然不会保存新建的表格
workbook.save('data.xlsx')
print("1-保存文件成功！")

# 如果user-agent 以字典键对形式作为headers的内容，就可以反爬成功，就不需要其他键对；
# 否则，需要加入headers下的更多键对形式。

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36'
}
# 这里输入需要爬取的总网络url  +'_'+str(1)
index = 'index'
all_url = 'http://wsjkw.nx.gov.cn/sy_279/yqfkdt/yqsd/'
# 定义一个条件退出标志
global conditon
""" ****************************************************第二步，从卫健委总网上爬取每天通报的分网站**************"""

def url_find(data):
    soup = bs(data, 'html.parser')
    # 查找关键字soup中的关键标签, 并选择性地过滤内容
    all_ul_lists = soup.find_all('ul', 'gl-list-l nx-bor2')
    for ui_list in all_ul_lists:
        twice_soup = bs(str(ui_list), 'html.parser')
        one_ul_lists = twice_soup.find_all('a')
        datas = twice_soup.find_all('span')
        for one_ul, ul_data in zip(one_ul_lists, datas):
            tag_content = one_ul.string
            if re.search('疫情情况', tag_content) is None:
                continue
            else:
                m_html = one_ul.get('href')
                # print(m_html)
                m_data = ul_data.string
                month = m_data[6:8]
                day = m_data[9:11]
                if int(month) <= 3 and int(day) < 32:
                    return -1
                else:
                    html_names.append(all_url+m_html[2:])
                    html_times.append(m_data[1:11])


print(" 正在查找网页...")
i = 0
while True:
    if i == 0:
        first_url = all_url
        # 请求网络，并添加头
        response_all = requests.get(first_url, headers=headers)
        # 设置编码格式，能正常显示中文
        response_all.encoding
        data = response_all.text
        i += 1
    else:
        other_url = all_url + 'index_' + str(i) + '.html'
        # print(other_url)
        response_all = requests.get(other_url, headers=headers)
        # 设置编码格式，能正常显示中文
        response_all.encoding
        data = response_all.text
        i += 1
    if url_find(data) == -1:
        break

""" ****************************************************第三步，从分网站里爬取有用信息，并保存一次数据**************"""
# print(html_names)
# print(html_times)
# 无新增
test1_html = 'http://wsjkw.nx.gov.cn/sy_279/yqfkdt/yqsd/202105/t20210528_2863377.html'
# 有新增
test2_html = 'http://wsjkw.nx.gov.cn/sy_279/yqfkdt/yqsd/202110/t20211022_3101803.html'


def choose_city(city, c_data, num):

    if city == '银川':
        Yinchuan[c_data] = num
    elif city == '固原':
        Guyuan[c_data] = num
    elif city == '石嘴山':
        Shizuishan[c_data] = num
    elif city == '吴忠':
        Wuzhong[c_data] = num
    elif city == '中卫':
        Zhongwei[c_data] = num


def config_peopel(c_html, c_data):
    m_total = 0
    response_one = requests.get(c_html, headers=headers)
    # 设置编码格式，能正常显示中文
    response_one.encoding
    peo_date = response_one.text
    soup = bs(peo_date, 'html.parser')
    # 查找关键字soup中的关键标签, 并选择性地过滤内容
    meta_content = soup.find_all(content=re.compile("疫情通报"))
    main_text = str(meta_content[0])
    for city in city_names:
        result = re.search(city, main_text)
        if result is None:
            choose_city(city, c_data, 0)
        else:
            position = result.span()[1]
            num = main_text[position]
            if num.isdigit():
                choose_city(city, c_data, num)  # 某某新增Num例
                m_total += int(num)
            elif num == '市':
                choose_city(city, c_data, 1)    # 某某市新增1例
                m_total += 1
            else:
                choose_city(city, c_data, 0)    # 其他病例，不录入统计
    if m_total != 0:
        Total[c_data] = m_total
    else:
        Total[c_data] = 0


print(" 网页查找完成！")
print(" 正在查找确诊人数...")
for html, data in zip(html_names, html_times):
    config_peopel(html, str(data))


i = 3
print(" 查找确诊人数完成！正在统计人数...")
for data, num in sorted(Total.items()):
        privacy.cell(i, 1).value = data
        privacy.cell(i-1, 2).value = num
        i += 1
i = 3
for data, num in sorted(Yinchuan.items()):
        city_sheet_yic.cell(i, 1).value = data
        city_sheet_yic.cell(i-1, 2).value = num
        i += 1
i = 3
for data, num in sorted(Guyuan.items()):
        city_sheet_guy.cell(i, 1).value = data
        city_sheet_guy.cell(i-1, 2).value = num
        i += 1
i = 3
for data, num in sorted(Shizuishan.items()):
        city_sheet_shz.cell(i, 1).value = data
        city_sheet_shz.cell(i-1, 2).value = num
        i += 1
i = 3
for data, num in sorted(Wuzhong.items()):
        city_sheet_wuz.cell(i, 1).value = data
        city_sheet_wuz.cell(i-1, 2).value = num
        i += 1
i = 3
for data, num in sorted(Zhongwei.items()):
        city_sheet_zhw.cell(i, 1).value = data
        city_sheet_zhw.cell(i-1, 2).value = num
        i += 1

print("2-统计人数完成！正在保存文件...")
# 这里需要保存一次，不然不会保存新建的表格
workbook.save('data.xlsx')
print("2-保存文件成功！")

"""****************************************************第三步，采集宁夏省市的迁移数据***************"""

"""
640000   宁夏省   
640100	 银川市	
640200	 石嘴山市									
640300	 吴忠市								
640400	 固原市									
640500	 中卫市	

"""
id_lists = [640000, 640100, 640200, 640300, 640400, 640500]
mig_types = {'move_in', 'move_out'}
time_slots = list(range(20210401, 20210431))+list(range(20210501, 20210532))+list(range(20210601, 20210631)) +\
             list(range(20210701, 20210732))+list(range(20210801, 20210832))+list(range(20210901, 20210931)) +\
             list(range(20211001, 20211032))+list(range(20211101, 20211123))
# 宁夏省的url
mig_url1 = 'http://huiyan.baidu.com/migration/historycurve.jsonp?dt=province&id='
mig_url2 = '&type='
migrate_headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 '
                          '(KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36'
}


def mig_city_choose(city_id, city_type, row, mig_value):
    if city_id == 640000:
        if city_type == 'move_in':
            privacy.cell(row, 3).value = mig_value
        else:
            privacy.cell(row, 4).value = mig_value
    elif city_id == 640100:
        if city_type == 'move_in':
            city_sheet_yic.cell(row, 3).value = mig_value
        else:
            city_sheet_yic.cell(row, 4).value = mig_value
    elif city_id == 640200:
        if city_type == 'move_in':
            city_sheet_shz.cell(row, 3).value = mig_value
        else:
            city_sheet_shz.cell(row, 4).value = mig_value
    elif city_id == 640300:
        if city_type == 'move_in':
            city_sheet_wuz.cell(row, 3).value = mig_value
        else:
            city_sheet_wuz.cell(row, 4).value = mig_value
    elif city_id == 640400:
        if city_type == 'move_in':
            city_sheet_guy.cell(row, 3).value = mig_value
        else:
            city_sheet_guy.cell(row, 4).value = mig_value
    elif city_id == 640500:
        if city_type == 'move_in':
            city_sheet_zhw.cell(row, 3).value = mig_value
        else:
            city_sheet_zhw.cell(row, 4).value = mig_value


for id in id_lists:
    for type in mig_types:
        mig_all_url = mig_url1 + str(id) + mig_url2 + type
        response_all = requests.get(mig_all_url, headers=migrate_headers)
        response_all.encoding
        pre_data = response_all.text
        # 截取内容的一部分
        after_data = pre_data[4:-1]
        # 要用json格式加载，不能直接加载使用文本
        data = json.loads(after_data)['data']
        all_list = data['list']
        i = 3
        for ind_date in time_slots:
            after_data = all_list[str(ind_date)]
            mig_city_choose(id, type, i, after_data)
            i += 1

print("3-正在保存文件...")
# 这里需要保存一次，不然不会保存新建的表格
workbook.save('data.xlsx')
print("3-保存文件成功！")


